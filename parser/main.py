import logging
import time
import pandas as pd
from selenium.webdriver.edge.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver import Edge, EdgeOptions
from selenium.common.exceptions import NoSuchElementException, StaleElementReferenceException
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.microsoft import EdgeChromiumDriverManager
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

# Настройка логгера
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

class ParsingWebsiteOstrovok:
    def __init__(self) -> None:
        options = EdgeOptions()
        options.add_argument("log-level=3")
        service = Service(EdgeChromiumDriverManager().install())
        self.driver = Edge(options=options, service=service)
        self.website = 'https://ostrovok.ru/hotel/russia'
        self.total_hotels_to_parse = 10  # Лимит на 10 отелей

    def forms_hotel_data(self, hotels_collected: int) -> list:
        hotels = []
        current_url = self.driver.current_url
        num_hotels_on_page = len(self.driver.find_elements(By.CLASS_NAME, 'hotel-wrapper'))

        for i in range(num_hotels_on_page):
            if hotels_collected >= self.total_hotels_to_parse:
                break
            
            attempts = 3
            while attempts > 0:
                try:
                    html_hotel_data = self.driver.find_elements(By.CLASS_NAME, 'hotel-wrapper')
                    data = html_hotel_data[i]
                    
                    name = data.find_element(By.CLASS_NAME, 'zen-hotelcard-name-link').text
                    hotel_url = data.find_element(By.CLASS_NAME, 'zen-hotelcard-name-link').get_attribute('href')
                    price = data.find_element(By.CLASS_NAME, 'zen-hotelcard-rate-price-value').text.replace(' ', '')
                    
                    reviews = self.get_reviews_from_hotel_page(hotel_url) if hotel_url else 'Отзывы отсутствуют'
                    
                    hotel_data = {
                        'Название': name,
                        'Цена': price,
                        'Отзывы': reviews,
                    }
                    hotels.append(hotel_data)
                    logging.info(f"Спарсенные данные отеля: {hotel_data}")
                    
                    # Сохранение информации об отеле в Excel после каждого успешного парсинга
                    save_hotel_data_to_excel(hotel_data)

                    self.driver.get(current_url)
                    time.sleep(2)

                    break
                except StaleElementReferenceException:
                    attempts -= 1
                    logging.warning("StaleElementReferenceException пойман, повторная попытка...")
                    time.sleep(1)
                
                if attempts == 0:
                    logging.error("Не удалось получить данные отеля после нескольких попыток.")
                    break
        
        return hotels

    def get_reviews_from_hotel_page(self, url: str) -> list[dict]:
        self.driver.get(url)
        WebDriverWait(self.driver, 10).until(
            EC.presence_of_all_elements_located((By.XPATH, "//*[contains(@class, 'Review_content__')]"))
        )

        reviews = []
        try:
            review_elements = self.driver.find_elements(By.XPATH, "//*[contains(@class, 'Review_content__kQpo_')]")
            for review in review_elements[:10]:  # Лимитируем до 10 отзывов
                try:
                    # Получаем оценку из TotalRating_content
                    try:
                        rating = review.find_element(By.XPATH, ".//span[contains(@class, 'TotalRating_content__k5u6S')]").text
                    except NoSuchElementException:
                        rating = "Оценка отсутствует"

                    # Получаем положительное описание отзыва
                    try:
                        positive_description = review.find_element(By.XPATH, ".//p[contains(@class, 'Review_plusTitle__')]/following-sibling::div/p").text
                    except NoSuchElementException:
                        positive_description = "Описание отсутствует"
                    
                    # Получаем отрицательное описание отзыва
                    try:
                        negative_description = review.find_element(By.XPATH, ".//p[contains(@class, 'Review_minusTitle__')]/following-sibling::div/p").text
                    except NoSuchElementException:
                        negative_description = "Описание отсутствует"
                    
                    # Добавляем отзыв в список
                    reviews.append({
                        'Оценка': rating,
                        'Что было хорошо': positive_description,
                        'Что было плохо': negative_description,
                    })
                except NoSuchElementException:
                    logging.warning("Оценка для отзыва не найдена, пропускаем отзыв.")
        except Exception as e:
            logging.error(f"Ошибка при сборе отзывов: {e}")

        return reviews

    def search_hotel_data(self) -> list:
        hotels = []
        page_number = 1
        hotels_collected = 0  # Счётчик отелей
        
        while hotels_collected < self.total_hotels_to_parse:
            url = f"{self.website}/?page={page_number}"
            logging.info(f"Открытие страницы: {url}")
            self.driver.get(url)
            
            WebDriverWait(self.driver, 10).until(EC.presence_of_all_elements_located((By.CLASS_NAME, 'hotel-wrapper')))
            
            hotels_batch = self.forms_hotel_data(hotels_collected)
            hotels.extend(hotels_batch)
            hotels_collected += len(hotels_batch)
            page_number += 1
            
            if hotels_collected >= self.total_hotels_to_parse:
                break
            
            if page_number > 10:
                logging.warning("Превышено число попыток загрузки страниц.")
                break

        return hotels[:self.total_hotels_to_parse]

    def close(self):
        self.driver.quit()

def save_hotel_data_to_excel(hotel_data):
    reviews_list = []
    
    for review in hotel_data['Отзывы']:
        reviews_list.append({
            'Отель': hotel_data['Название'],
            'Цена': hotel_data['Цена'],
            'Оценка': review['Оценка'],
            'Что было хорошо': review['Что было хорошо'],
            'Что было плохо': review['Что было плохо']
        })
    
    # Преобразуем данные в DataFrame
    df = pd.DataFrame(reviews_list)
    
    # Открываем или создаем файл Excel
    try:
        with pd.ExcelWriter('hotel_reviews.xlsx', engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
            df.to_excel(writer, index=False, header=False, startrow=writer.sheets['Отзывы'].max_row, sheet_name='Отзывы')
    except FileNotFoundError:
        # Если файл не существует, создаем его с заголовками
        with pd.ExcelWriter('hotel_reviews.xlsx', engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Отзывы')
            sheet = writer.sheets['Отзывы']
            
            # Применяем стиль заголовков
            header_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            for cell in sheet[1]:
                cell.fill = header_fill
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')

if __name__ == "__main__":
    parser = ParsingWebsiteOstrovok()
    try:
        hotels_data = parser.search_hotel_data()
        logging.info(f"Всего спарсено отелей: {len(hotels_data)}")
    finally:
        parser.close()
